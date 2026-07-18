import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

SHEET_NAMES = {
    "market_index": "大盤指數",
    "market_chip": "大盤三大法人",
    "market_margin": "全市場融資融券",
    "stock_chip": "個股三大法人",
    "margin_balance": "個股融資融券",
    "stock_quote": "個股收盤價",
    "foreign_futures_oi": "外資期貨未平倉",
    "monthly_revenue": "月營收動能",
    "financial_income": "財報毛利率",
    "stock_price_action": "個股價格結構",
}

COLUMN_NAMES = {
    "trade_date": "交易日期",
    "stock_id": "股票代號",
    "stock_name": "股票名稱",
    "close": "收盤價",
    "change_pts": "漲跌點數",
    "change_pct": "漲跌幅(%)",
    "volume_yi": "成交量(億元)",
    "foreign_net": "外資買賣超",
    "trust_net": "投信買賣超",
    "dealer_self_net": "自營商自行買賣超",
    "dealer_hedge_net": "自營商避險買賣超",
    "dealer_total_net": "自營商合計買賣超",
    "margin_balance": "融資餘額(張)",
    "margin_balance_chg": "融資餘額增減(張)",
    "short_balance": "融券餘額(張)",
    "short_balance_chg": "融券餘額增減(張)",
    "margin_balance_lots": "全市場融資餘額(張)",
    "margin_balance_lots_chg": "全市場融資餘額增減(張)",
    "short_balance_lots": "全市場融券餘額(張)",
    "short_balance_lots_chg": "全市場融券餘額增減(張)",
    "margin_balance_yi": "全市場融資餘額(億元)",
    "margin_balance_yi_chg": "全市場融資餘額增減(億元)",
    "dividend_yield": "殖利率(%)",
    "pe": "本益比",
    "pb": "股價淨值比",
    "contract_code": "契約名稱",
    "institution_type": "法人類別",
    "oi_long": "多方未平倉",
    "oi_short": "空方未平倉",
    "oi_net": "淨未平倉",
    "period": "期別",
    "industry": "產業別",
    "revenue": "營業收入(仟元)",
    "revenue_prev_month": "上月營收(仟元)",
    "revenue_last_year_month": "去年當月營收(仟元)",
    "revenue_mom_pct": "營收月增率(%)",
    "revenue_yoy_pct": "營收年增率(%)",
    "report_date": "公告日期",
    "fiscal_year": "年度",
    "fiscal_quarter": "季別",
    "cost": "營業成本(仟元)",
    "gross_profit": "營業毛利(仟元)",
    "gross_margin": "毛利率(%)",
    "operating_income": "營業利益(仟元)",
    "operating_margin": "營益率(%)",
    "net_income": "本期淨利(仟元)",
    "net_margin": "淨利率(%)",
    "eps": "每股盈餘(元)",
    "open": "開盤價",
    "high": "最高價",
    "low": "最低價",
    "volume_lots": "成交量(張)",
    "turnover_yi": "成交金額(億元)",
}


def export_excel(db_path, out_path):
    """把資料庫裡每張表匯出成 Excel 的一個分頁,最新日期排最上面。"""
    conn = sqlite3.connect(db_path)
    tables = [r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")]

    wb = Workbook()
    wb.remove(wb.active)

    for t in tables:
        cols = [r[1] for r in conn.execute(f"PRAGMA table_info({t})")]
        order_col = "trade_date" if "trade_date" in cols else cols[0]
        rows = conn.execute(f"SELECT * FROM {t} ORDER BY {order_col} DESC").fetchall()

        ws = wb.create_sheet(title=SHEET_NAMES.get(t, t)[:31])
        headers = [COLUMN_NAMES.get(c, c) for c in cols]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

        widths = [len(c) for c in headers]
        for row in rows:
            ws.append(list(row))
            for i, v in enumerate(row):
                widths[i] = max(widths[i], len(str(v)))
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(w + 2, 40)

    conn.close()
    wb.save(out_path)
    return out_path
